from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config.settings import settings
from app.schemas.ai_request import AiRequest


def load_payment_slots_from_excel(req: AiRequest) -> dict[str, Any]:
    xlsx = _find_latest_xlsx(Path(settings.payment_input_dir))
    if xlsx is None:
        return {}
    slots = _parse_order_form(xlsx)
    if not slots:
        return {}

    requested_payment_id = _to_text(req.get_slot("payment_id"))
    loaded_payment_id = _to_text(slots.get("payment_id"))
    if requested_payment_id and loaded_payment_id and requested_payment_id != loaded_payment_id:
        return {}
    return slots


def _find_latest_xlsx(base_dir: Path) -> Path | None:
    if not base_dir.exists() or not base_dir.is_dir():
        return None
    files = [p for p in base_dir.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")]
    if not files:
        return None
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0]


def _parse_order_form(path: Path) -> dict[str, Any]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = _select_best_sheet(wb.worksheets)
        if ws is None:
            return {}

        issue_date = _extract_value_after_label(ws, {"발주일", "발주일자", "발주일", "발 주 일"})
        total_price = _extract_numeric_after_label(ws, {"발주금액", "총액", "합계금액", "합계"})
        payment_id = _extract_order_no(ws)
        supplier_name = _extract_value_after_label(ws, {"상호명", "상 호 명", "외주처"})
        supplier_contact = _extract_value_after_label(ws, {"연락처", "전화번호", "연 락 처"})
        items = _extract_items(ws)

        paid_at = _to_iso_date_text(issue_date)
        month = paid_at[:7] if paid_at and len(paid_at) >= 7 else None
        total = _to_int(total_price)
        if total is None and items:
            total = sum(
                _to_int(item.get("supply_amount")) or _to_int(item.get("amount")) or 0
                for item in items
            )
            if total == 0:
                total = None

        slots: dict[str, Any] = {
            "payment_id": payment_id or None,
            "paid_at": paid_at,
            "month": month,
            "billing_month": month,
            "total_price": total,
            "target_type": "order_request_form",
            "supplier_name": supplier_name or None,
            "supplier_contact": supplier_contact or None,
            "items": items or None,
            "source_file": path.name,
        }
        return {k: v for k, v in slots.items() if v not in (None, "", [])}
    finally:
        wb.close()


def _select_best_sheet(sheets: list[Worksheet]) -> Worksheet | None:
    if not sheets:
        return None

    keywords = {"발주", "품명", "수량", "단가"}
    best: Worksheet | None = None
    best_score = -1
    for ws in sheets:
        score = 0
        for row in ws.iter_rows(min_row=1, max_row=40, max_col=40, values_only=True):
            for cell in row:
                norm = _norm_korean(cell)
                if not norm:
                    continue
                if any(k in norm for k in keywords):
                    score += 1
        if score > best_score:
            best = ws
            best_score = score
    return best


def _extract_order_no(ws: Worksheet) -> str:
    for row in ws.iter_rows(min_row=1, max_row=12, min_col=1, max_col=40, values_only=False):
        for idx, cell in enumerate(row):
            if _norm_korean(cell.value) not in {"no", "no."}:
                continue
            for step in range(1, 6):
                if idx + step >= len(row):
                    break
                value = _to_text(row[idx + step].value)
                if value and _norm_korean(value) not in {"no", "no."}:
                    return value
    return ""


def _extract_value_after_label(ws: Worksheet, labels: set[str]) -> str:
    normalized_labels = {_norm_korean(v) for v in labels}
    for row in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=45, values_only=False):
        for idx, cell in enumerate(row):
            if _norm_korean(cell.value) not in normalized_labels:
                continue
            for step in range(1, 12):
                if idx + step >= len(row):
                    break
                value = _to_text(row[idx + step].value)
                if value and _norm_korean(value) not in normalized_labels:
                    return value
    return ""


def _extract_numeric_after_label(ws: Worksheet, labels: set[str]) -> int | None:
    text = _extract_value_after_label(ws, labels)
    if text:
        parsed = _to_int(text)
        if parsed is not None:
            return parsed

    normalized_labels = {_norm_korean(v) for v in labels}
    for row in ws.iter_rows(min_row=1, max_row=70, min_col=1, max_col=45, values_only=False):
        for idx, cell in enumerate(row):
            if _norm_korean(cell.value) not in normalized_labels:
                continue
            for step in range(1, 12):
                if idx + step >= len(row):
                    break
                parsed = _to_int(row[idx + step].value)
                if parsed is not None:
                    return parsed
    return None


def _extract_items(ws: Worksheet) -> list[dict[str, Any]]:
    header_row, col_map = _find_item_header(ws)
    if header_row == 0 or not col_map:
        return []

    items: list[dict[str, Any]] = []
    empty_streak = 0
    for r in range(header_row + 1, header_row + 80):
        if r > ws.max_row:
            break

        name = _cell_text(ws, r, col_map.get("name"))
        spec = _cell_text(ws, r, col_map.get("spec"))
        qty = _to_int(_cell_value(ws, r, col_map.get("qty")))
        unit_price = _to_int(_cell_value(ws, r, col_map.get("unit_price")))
        amount = _to_int(_cell_value(ws, r, col_map.get("amount")))
        supply_amount = _to_int(_cell_value(ws, r, col_map.get("supply_amount")))
        tax_amount = _to_int(_cell_value(ws, r, col_map.get("tax_amount")))
        note = _cell_text(ws, r, col_map.get("note"))

        first_col_value = _cell_text(ws, r, col_map.get("no"))
        if _norm_korean(first_col_value).startswith("합계") or _norm_korean(first_col_value).startswith("1납기"):
            break

        is_empty = not any([name, spec, qty, unit_price, amount, supply_amount, tax_amount, note])
        if is_empty:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0

        item = {
            "prod_nm": name or None,
            "spec": spec or None,
            "qty": qty,
            "unit_price": unit_price,
            "amount": amount,
            "supply_amount": supply_amount,
            "tax_amount": tax_amount,
            "note": note or None,
        }
        items.append({k: v for k, v in item.items() if v not in (None, "", [])})
    return items


def _find_item_header(ws: Worksheet) -> tuple[int, dict[str, int]]:
    for r in range(1, min(ws.max_row, 90) + 1):
        cells = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 60) + 1)]
        normalized = [_norm_korean(v) for v in cells]
        if "품명" not in normalized:
            continue
        if "수량" not in normalized and "단가" not in normalized:
            continue
        col_map: dict[str, int] = {}
        for c, text in enumerate(normalized, start=1):
            if text in {"no", "no.", "번호"}:
                col_map["no"] = c
            elif text == "품명":
                col_map["name"] = c
            elif text in {"규격"}:
                col_map["spec"] = c
            elif text in {"수량"}:
                col_map["qty"] = c
            elif text in {"단가"}:
                col_map["unit_price"] = c
            elif text in {"금액"}:
                col_map["amount"] = c
            elif text in {"공급가액", "공급가"}:
                col_map["supply_amount"] = c
            elif text in {"세액"}:
                col_map["tax_amount"] = c
            elif text in {"비고"}:
                col_map["note"] = c
        return r, col_map
    return 0, {}


def _cell_value(ws: Worksheet, row: int, col: int | None) -> Any:
    if not col:
        return None
    return ws.cell(row=row, column=col).value


def _cell_text(ws: Worksheet, row: int, col: int | None) -> str:
    return _to_text(_cell_value(ws, row, col))


def _norm_korean(value: Any) -> str:
    text = _to_text(value)
    if not text:
        return ""
    text = text.lower()
    text = re.sub(r"[\s\r\n\t]+", "", text)
    text = text.replace(":", "").replace("：", "")
    return text


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None
    text = _to_text(value)
    if not text:
        return None
    text = re.sub(r"[^0-9.-]", "", text)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _to_iso_date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _to_text(value)
    if not text:
        return None

    # yyyy-mm-dd or yyyy.mm.dd or yyyy/mm/dd
    match = re.search(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"

    # yyyymmdd
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"
    return text


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
