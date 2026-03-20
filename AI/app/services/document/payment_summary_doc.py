from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config.settings import settings
from app.schemas.ai_request import AiRequest
from app.services.actions.event_sink import publish_action_event
from app.services.document.backend_file_uploader import upload_generated_file
from app.services.document.draft_writer import write_document_draft
from app.services.document.payment_excel_loader import load_payment_slots_from_excel


def make_payment_summary(req: AiRequest) -> tuple[str, dict]:
    _enrich_slots_from_excel(req)
    if req.intent == "PAYMENT_STATUS_SUMMARY":
        answer, payload = _build_payment_status_summary(req)
        draft_path = write_document_draft("payment_status_summary", payload=payload, user_id=req.user_id)
        event = publish_action_event(
            "payment_status_summary_created",
            {"user_id": req.user_id, "draft_path": draft_path, "summary": payload},
        )
        return answer, {"draft_path": draft_path, "draft_type": "payment_status_summary", "event_status": event}

    answer, payload = _build_payment_document_summary(req)

    xlsx_fname = None
    file_id = None
    try:
        xlsx_path = _write_summary_xlsx(payload, user_id=req.user_id or "admin")
        building_id = _to_int(req.get_slot("building_id"))
        try:
            uploaded = upload_generated_file(
                xlsx_path,
                file_parent_type="AI_DOCUMENT",
                file_parent_id=building_id if building_id is not None else 0,
            )
        finally:
            xlsx_path.unlink(missing_ok=True)

        xlsx_fname = uploaded["file_name"]
        file_id = uploaded["file_id"]
        payload["file_id"] = file_id
        payload["xlsx_file_name"] = xlsx_fname
        payload["download_url"] = uploaded["download_url"]
        payload["view_url"] = uploaded["view_url"]
    except Exception as _e:
        import logging
        logging.getLogger(__name__).warning("xlsx 생성 실패: %s", _e)
        xlsx_fname = None

    draft_path = write_document_draft("payment_summary_document", payload=payload, user_id=req.user_id)
    event = publish_action_event(
        "payment_summary_document_created",
        {"user_id": req.user_id, "draft_path": draft_path, "summary": payload},
    )
    return answer, {
        "draft_path": draft_path,
        "draft_type": "payment_summary_document",
        "event_status": event,
        "file_id": file_id,
        "xlsx_file_name": xlsx_fname,
        "download_url": payload.get("download_url"),
    }


def _write_summary_xlsx(payload: dict[str, Any], user_id: str = "admin") -> Path:
    out = Path(settings.payment_order_output_dir)
    out.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%dT%H%M%SZ")
    safe_user = _safe(user_id)
    month_tag = str(payload.get("month") or payload.get("billing_month") or "").replace("-", "")
    fname = f"payment_summary_{month_tag}_{safe_user}_{ts}_{uuid4().hex[:6]}.xlsx"
    fpath = out / fname

    wb = Workbook()
    ws = wb.active
    ws.title = "결제 요약"

    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    def hdr(c, v):
        c.value = v; c.font = HDR_FONT
        c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER

    def cel(c, v, bold=False, align=None):
        c.value = v; c.font = Font(bold=bold, size=10)
        c.alignment = align or LEFT; c.border = BORDER

    ws.merge_cells("A1:H1")
    ws["A1"].value = "결제 내역 요약서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  담당: {user_id}"
    ws["A2"].alignment = CENTER
    ws["A2"].font = Font(size=9, color="666666")

    ws.append([])

    ws.append(["항목", "내용"])
    hdr(ws.cell(ws.max_row, 1), "항목")
    hdr(ws.cell(ws.max_row, 2), "내용")

    total_price = payload.get("total_price")
    for label, value in [
        ("정산 월",    payload.get("month") or payload.get("billing_month") or "-"),
        ("결제 ID",    payload.get("payment_id") or "-"),
        ("결제 일자",  payload.get("paid_at") or "-"),
        ("총 결제금액", f"{total_price:,}원" if total_price else "-"),
        ("공급업체",   payload.get("supplier_name") or "-"),
        ("연락처",     payload.get("supplier_contact") or "-"),
        ("문서 유형",  payload.get("target_type") or "-"),
        ("원본 파일",  payload.get("source_file") or "-"),
    ]:
        ws.append([label, str(value)])
        r = ws.max_row
        cel(ws.cell(r, 1), label, bold=True)
        cel(ws.cell(r, 2), str(value))

    ws.append([])

    items = payload.get("items") or []
    if items:
        headers = ["No.", "품명", "규격", "수량", "단가", "공급가액", "세액", "비고"]
        ws.append(headers)
        hr = ws.max_row
        for ci, h in enumerate(headers, 1):
            hdr(ws.cell(hr, ci), h)

        total_supply = 0
        for idx, item in enumerate(items, 1):
            supply = item.get("supply_amount") or item.get("amount") or 0
            tax    = item.get("tax_amount") or 0
            total_supply += supply or 0
            row_data = [
                idx,
                item.get("prod_nm") or "",
                item.get("spec") or "",
                item.get("qty") or "",
                f"{item['unit_price']:,}" if item.get("unit_price") else "",
                f"{supply:,}" if supply else "",
                f"{tax:,}" if tax else "",
                item.get("note") or "",
            ]
            ws.append(row_data)
            r = ws.max_row
            for ci, v in enumerate(row_data, 1):
                cel(ws.cell(r, ci), v, align=CENTER if ci in (1, 4) else LEFT)

        ws.append(["", "합   계", "", "", "", f"{total_supply:,}", "", ""])
        r = ws.max_row
        for ci in range(1, 9):
            c = ws.cell(r, ci)
            c.font = Font(bold=True, size=10); c.border = BORDER
            c.fill = TOTAL_FILL
            c.alignment = CENTER if ci != 2 else LEFT

    for i, w in enumerate([6, 22, 12, 6, 12, 14, 12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A4"

    wb.save(fpath)
    return fpath


def _enrich_slots_from_excel(req: AiRequest) -> None:
    loaded = load_payment_slots_from_excel(req)
    if not loaded:
        return
    for key, value in loaded.items():
        if req.get_slot(key) in (None, "") and value not in (None, ""):
            req.slots[key] = value


def _build_payment_document_summary(req: AiRequest) -> tuple[str, dict]:
    month       = str(req.get_slot("month") or req.get_slot("billing_month") or "N/A")
    payment_id  = str(req.get_slot("payment_id") or "N/A")
    total_price = _to_int(req.get_slot("total_price"))
    paid_at     = str(req.get_slot("paid_at") or "unpaid")
    target_type = str(req.get_slot("target_type") or "general")
    supplier    = str(req.get_slot("supplier_name") or "")
    items       = req.get_slot("items") or []

    price_text = f"{total_price:,}원" if total_price is not None else "N/A"
    answer = (
        f"결제 요약 문서가 준비되었습니다.\n"
        f"📅 정산월: {month} | 🔖 결제ID: {payment_id}\n"
        f"💰 금액: {price_text} | 📆 결제일: {paid_at}"
        + (f" | 🏢 공급업체: {supplier}" if supplier else "")
        + (f"\n📦 품목 수: {len(items)}건" if items else "")
    )
    payload = {
        "month": month, "payment_id": payment_id,
        "total_price": total_price, "paid_at": paid_at,
        "target_type": target_type,
        "supplier_name": req.get_slot("supplier_name"),
        "supplier_contact": req.get_slot("supplier_contact"),
        "source_file": req.get_slot("source_file"),
        "items": items if isinstance(items, list) else [],
    }
    return answer, payload


def _build_payment_status_summary(req: AiRequest) -> tuple[str, dict]:
    billing_month = str(req.get_slot("billing_month") or req.get_slot("month") or "N/A")
    payment_st    = str(req.get_slot("payment_st") or "UNKNOWN").upper()
    charge_status = str(req.get_slot("charge_status") or "UNKNOWN").upper()
    due_date_raw  = req.get_slot("due_date")
    due_date = _to_date(due_date_raw)
    due_text = "N/A"
    if due_date is not None:
        days_left = (due_date - date.today()).days
        due_text = (f"연체 {-days_left}일" if days_left < 0
                    else "오늘 마감" if days_left == 0 else f"D-{days_left}")
    answer = (f"결제 현황 요약: 정산월={billing_month}, 결제상태={payment_st}, "
              f"청구상태={charge_status}, 마감={due_text}.")
    return answer, {"billing_month": billing_month, "payment_status": payment_st,
                    "charge_status": charge_status, "due_text": due_text}


def _safe(value: str) -> str:
    text = "".join(c if c.isalnum() or c in "-_" else "_" for c in (value or "").lower().strip())
    return text.strip("_") or "unknown"


def _to_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _to_date(value: object) -> date | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError:
        return None
