from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config.settings import settings
from app.schemas.ai_request import AiRequest
from app.services.actions.event_sink import publish_action_event
from app.services.tools.tool_executor import execute_tool

logger = logging.getLogger(__name__)


def _build_download_url(file_name: str) -> str:
    return f"/ai/payment/order-form/download/{file_name}"

# ════════════════════════════════════════════════════════════════════════════
# 빌딩별 결제 내역 리포트 (DB → xlsx)
# ════════════════════════════════════════════════════════════════════════════

# 스타일 상수
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_TOT_FILL = PatternFill("solid", fgColor="D9E1F2")
_TOT_FONT = Font(bold=True, size=10)
_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")


def generate_billing_report(
    admin_id: str,
    target_month: str | None = None,
    building_nm: str | None = None,
) -> dict[str, Any]:
    """
    빌딩별 결제 내역 xlsx 리포트 생성.
    target_month : 'YYYY-MM' (None → 이번달)
    building_nm  : 특정 빌딩명 (None → 전체)
    반환: {file_name, download_url, summary 통계}
    """
    month = target_month or datetime.now().strftime("%Y-%m")
    bld_filter = f"AND b.building_nm LIKE '%{building_nm}%'" if building_nm else ""

    # ── 1) monthly_charge (월세·관리비 청구) ────────────────────────────────
    sql_mc = f"""
SELECT b.building_nm, r.room_no, u.user_nm, u.user_id,
       mc.billing_dt, mc.charge_type, mc.price, mc.charge_st
FROM monthly_charge mc
JOIN contract c ON mc.contract_id = c.contract_id
JOIN rooms r    ON c.room_id = r.room_id
JOIN building b ON r.building_id = b.building_id
JOIN users u    ON c.user_id = u.user_id
WHERE mc.billing_dt = '{month}' AND b.delete_yn = 'N' {bld_filter}
ORDER BY b.building_nm, r.room_no, mc.charge_type
LIMIT 500""".strip()

    # ── 2) payment (룸서비스 등 기타 결제) ─────────────────────────────────
    sql_pay = f"""
SELECT b.building_nm, r.room_no, u.user_nm, p.payment_id,
       DATE_FORMAT(p.paid_at, '%Y-%m-%d') AS paid_date,
       sg.service_goods_nm, p.total_price, p.captured_price,
       p.payment_st, p.provider
FROM payment p
JOIN contract c       ON c.user_id = p.user_id AND c.contract_st = 'active'
JOIN rooms r          ON c.room_id = r.room_id
JOIN building b       ON r.building_id = b.building_id
JOIN users u          ON p.user_id = u.user_id
LEFT JOIN service_goods sg ON p.service_goods_id = sg.service_goods_id
WHERE DATE_FORMAT(p.paid_at, '%Y-%m') = '{month}'
  AND p.payment_st = 'paid' AND b.delete_yn = 'N' {bld_filter}
ORDER BY b.building_nm, r.room_no, p.paid_at
LIMIT 500""".strip()

    mc_res  = execute_tool("query_database_admin", {"sql": sql_mc,  "description": "월세·관리비 조회"}, admin_id)
    pay_res = execute_tool("query_database_admin", {"sql": sql_pay, "description": "기타결제 조회"},   admin_id)

    mc_rows  = mc_res.get("data")  or [] if mc_res.get("success")  else []
    pay_rows = pay_res.get("data") or [] if pay_res.get("success") else []
    logger.info("[BillingReport] monthly_charge=%d건, payment=%d건", len(mc_rows), len(pay_rows))

    # ── xlsx 생성 ─────────────────────────────────────────────────────────
    fname = _billing_file_name(month, building_nm, admin_id)
    out   = Path(settings.payment_order_output_dir)
    out.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, mc_rows, pay_rows, month)
    _sheet_monthly_charge(wb, mc_rows, month)
    _sheet_payment(wb, pay_rows, month)

    save_path = out / fname
    wb.save(save_path)

    file_size = save_path.stat().st_size if save_path.exists() else 0
    logger.info("[BillingReport] 파일 저장 완료: %s (%d bytes)", save_path, file_size)
    if file_size == 0:
        raise RuntimeError(f"저장된 파일 크기가 0입니다: {save_path}")

    mc_total  = sum(int(r.get("price", 0) or 0) for r in mc_rows)
    pay_total = sum(int(r.get("captured_price", 0) or 0) for r in pay_rows)
    buildings = sorted({r.get("building_nm", "") for r in mc_rows + pay_rows} - {""})

    return {
        "month":        month,
        "buildings":    buildings,
        "mc_count":     len(mc_rows),
        "mc_total":     mc_total,
        "pay_count":    len(pay_rows),
        "pay_total":    pay_total,
        "grand_total":  mc_total + pay_total,
        "file_name":    fname,
        "download_url": _build_download_url(fname),
    }


# ── 시트 작성 ────────────────────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, mc_rows: list, pay_rows: list, month: str) -> None:
    """빌딩별 × 건별 전체 상세 내역 시트 (첫 번째)."""
    ws = wb.create_sheet("전체 상세 내역", 0)
    _title(ws, f"빌딩별 결제 상세 내역  ({month})", 9)
    _hdr_row(ws, ["건물명", "호수", "입주자", "일자", "구분", "유형/상품", "금액", "납부상태", "비고"])

    ct_map = {"rent": "월세", "manage_fee": "관리비", "deposit": "보증금"}
    cs_map = {"paid": "✅ 납부", "unpaid": "❌ 미납", "pending": "⏳ 대기"}

    buildings = sorted({r.get("building_nm", "") for r in mc_rows + pay_rows} - {""})
    grand_total = 0

    for bld in buildings:
        # 해당 빌딩 월세·관리비
        mc_b  = [r for r in mc_rows  if r.get("building_nm") == bld]
        pay_b = [r for r in pay_rows if r.get("building_nm") == bld]

        if not mc_b and not pay_b:
            continue

        # 빌딩 소계 행 (구분선)
        ws.append([bld, "", "", "", "", "", "", "", ""])
        r = ws.max_row
        for ci in range(1, 10):
            c = ws.cell(r, ci)
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2E75B6")
            c.border = _BORDER
            c.alignment = _LEFT
        ws.row_dimensions[r].height = 18

        bld_total = 0

        for row in mc_b:
            price = int(row.get("price") or 0)
            bld_total += price
            _data_row(ws, [
                row.get("building_nm", ""),
                row.get("room_no", ""),
                row.get("user_nm", ""),
                row.get("billing_dt", ""),
                "월세·관리비",
                ct_map.get(row.get("charge_type", ""), row.get("charge_type", "")),
                price,
                cs_map.get(row.get("charge_st", ""), row.get("charge_st", "")),
                "",
            ], 9)

        for row in pay_b:
            captured = int(row.get("captured_price") or 0)
            bld_total += captured
            _data_row(ws, [
                row.get("building_nm", ""),
                row.get("room_no", ""),
                row.get("user_nm", ""),
                row.get("paid_date", ""),
                "기타결제",
                row.get("service_goods_nm", "-"),
                captured,
                "✅ 결제완료",
                row.get("provider", ""),
            ], 9)

        # 빌딩 소계
        ws.append(["", "", "", "", bld + " 소계", "", bld_total, "", ""])
        r = ws.max_row
        for ci in range(1, 10):
            c = ws.cell(r, ci)
            c.font = _TOT_FONT; c.fill = _TOT_FILL; c.border = _BORDER
            if isinstance(c.value, int):
                c.number_format = "#,##0"; c.alignment = _RIGHT
            else:
                c.alignment = _LEFT
        grand_total += bld_total

    # 총합계
    ws.append(["", "", "", "", "총 합계", "", grand_total, "", ""])
    r = ws.max_row
    for ci in range(1, 10):
        c = ws.cell(r, ci)
        c.font = Font(bold=True, size=11); c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.border = _BORDER
        if isinstance(c.value, int):
            c.number_format = "#,##0"; c.alignment = _RIGHT
        else:
            c.alignment = _LEFT

    _col_widths(ws, [16, 8, 12, 12, 12, 18, 14, 12, 12])
    ws.freeze_panes = "A3"


def _sheet_monthly_charge(wb: Workbook, rows: list, month: str) -> None:
    ws = wb.create_sheet("월세·관리비 청구")
    _title(ws, f"월세·관리비 청구 내역  ({month})", 8)
    _hdr_row(ws, ["건물명", "호수", "입주자", "청구월", "청구유형", "금액", "납부상태", "비고"])

    ct_map = {"rent": "월세", "manage_fee": "관리비", "deposit": "보증금"}
    cs_map = {"paid": "✅ 납부", "unpaid": "❌ 미납", "pending": "⏳ 대기"}
    total = 0
    for r in rows:
        price = int(r.get("price") or 0)
        total += price
        _data_row(ws, [
            r.get("building_nm", ""), r.get("room_no", ""), r.get("user_nm", ""),
            r.get("billing_dt", ""),
            ct_map.get(r.get("charge_type", ""), r.get("charge_type", "")),
            price,
            cs_map.get(r.get("charge_st", ""), r.get("charge_st", "")),
            "",
        ], 8)

    _total_row(ws, ["", "합  계", "", "", "", total, "", ""], 8)
    _col_widths(ws, [16, 8, 12, 12, 12, 14, 12, 10])


def _sheet_payment(wb: Workbook, rows: list, month: str) -> None:
    ws = wb.create_sheet("기타 결제")
    _title(ws, f"기타 결제 내역 (룸서비스 등)  ({month})", 8)
    _hdr_row(ws, ["건물명", "호수", "입주자", "결제일", "상품명", "결제금액", "실결제", "결제수단"])

    total = 0
    for r in rows:
        captured = int(r.get("captured_price") or 0)
        total += captured
        _data_row(ws, [
            r.get("building_nm", ""), r.get("room_no", ""), r.get("user_nm", ""),
            r.get("paid_date", ""), r.get("service_goods_nm", "-"),
            int(r.get("total_price") or 0), captured, r.get("provider", ""),
        ], 8)

    _total_row(ws, ["", "합  계", "", "", "", "", total, ""], 8)
    _col_widths(ws, [16, 8, 12, 12, 18, 14, 14, 12])


# ── 공통 스타일 헬퍼 ──────────────────────────────────────────────────────────

def _title(ws, text: str, cols: int) -> None:
    last = get_column_letter(cols)
    ws.merge_cells(f"A1:{last}1")
    ws["A1"].value = text
    ws["A1"].font  = Font(bold=True, size=13)
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 26


def _hdr_row(ws, headers: list) -> None:
    ws.append(headers)
    r = ws.max_row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(r, ci)
        c.value = h; c.font = _HDR_FONT
        c.fill = _HDR_FILL; c.alignment = _CENTER; c.border = _BORDER
    ws.row_dimensions[r].height = 20


def _data_row(ws, values: list, col_count: int) -> None:
    ws.append(values)
    r = ws.max_row
    for ci in range(1, col_count + 1):
        c = ws.cell(r, ci)
        c.font = Font(size=10); c.border = _BORDER
        if isinstance(c.value, int) and ci > 1:
            c.number_format = "#,##0"; c.alignment = _RIGHT
        else:
            c.alignment = _CENTER if ci == 2 else _LEFT


def _total_row(ws, values: list, col_count: int) -> None:
    ws.append(values)
    r = ws.max_row
    for ci in range(1, col_count + 1):
        c = ws.cell(r, ci)
        c.font = _TOT_FONT; c.fill = _TOT_FILL; c.border = _BORDER
        if isinstance(c.value, int) and c.value > 0:
            c.number_format = "#,##0"; c.alignment = _RIGHT
        else:
            c.alignment = _LEFT


def _col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _billing_file_name(month: str, building_nm: str | None, user_id: str) -> str:
    ts    = datetime.now().strftime("%Y%m%dT%H%M%SZ")
    m_tag = month.replace("-", "")
    b_tag = _safe("".join(c if c.isalnum() else "_" for c in (building_nm or "all")))
    u_tag = _safe(user_id or "admin")
    return f"billing_report_{m_tag}_{b_tag}_{u_tag}_{uuid4().hex[:6]}.xlsx"


def create_order_form_from_suggestion(req: AiRequest) -> tuple[str, dict[str, Any]]:
    if req.get_slot("approved") is False:
        return "Order form generation skipped because admin approval is false.", {"approved": False}

    items = _collect_items(req)
    if not items:
        return "No approved items were provided for order form generation.", {"approved": True, "items": []}

    template_path = Path(settings.payment_order_template_path).resolve()
    if not template_path.exists():
        return f"Order template not found: {template_path}", {"approved": True, "items": items}

    output_path = _build_output_path(req)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    try:
        ws = _select_sheet(wb.worksheets)
        _fill_header(ws, req)
        _fill_items(ws, items)
        wb.save(output_path)
    finally:
        wb.close()

    payload = {
        "approved": True,
        "building_id": _to_int(req.get_slot("building_id")),
        "month": str(req.get_slot("month") or req.get_slot("billing_month") or ""),
        "item_count": len(items),
        "order_form_path": str(output_path),
        "file_name": output_path.name,
    }
    event = publish_action_event("payment_order_form_created", {"user_id": req.user_id, **payload})
    payload["event_status"] = event
    return f"Order form generated: {output_path.name}", payload


def _build_output_path(req: AiRequest) -> Path:
    base = Path(settings.payment_order_output_dir)
    ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    user = _safe(req.user_id or "admin")
    building = _safe(str(req.get_slot("building_id") or "all"))
    name = f"purchase_order_{building}_{user}_{ts}_{uuid4().hex[:8]}.xlsx"
    return base / name


def _select_sheet(sheets: list[Worksheet]) -> Worksheet:
    if not sheets:
        raise ValueError("No worksheet in template")
    best = sheets[0]
    best_score = -1
    for ws in sheets:
        score = 0
        for row in ws.iter_rows(min_row=1, max_row=25, max_col=40, values_only=True):
            for value in row:
                text = str(value or "").replace(" ", "")
                if "발주서" in text:
                    score += 2
                if text in {"품명", "수량", "단가", "공급가액"}:
                    score += 1
        if score > best_score:
            best = ws
            best_score = score
    return best


def _fill_header(ws: Worksheet, req: AiRequest) -> None:
    today = str(req.get_slot("order_date") or datetime.now().date().isoformat())
    order_no = str(req.get_slot("order_no") or req.get_slot("payment_id") or uuid4().hex[:8].upper())
    buyer_name = str(req.get_slot("buyer_name") or "유니플레이스")
    buyer_tel = str(req.get_slot("buyer_tel") or req.get_slot("contact") or "")
    supplier_name = str(req.get_slot("supplier_name") or req.get_slot("affiliate_name") or "")
    supplier_tel = str(req.get_slot("supplier_contact") or req.get_slot("supplier_tel") or "")

    _put_after_label(ws, {"발 주 일", "발주일", "발주일자"}, today)
    _put_after_label(ws, {"NO.", "NO", "NO. "}, order_no)
    _put_after_label(ws, {"상 호 명", "상호명", "외주처"}, supplier_name)
    _put_after_label(ws, {"연 락 처", "연락처", "전화번호"}, supplier_tel or buyer_tel)
    _put_after_label(ws, {"발 주 자", "발주자"}, buyer_name)


def _fill_items(ws: Worksheet, items: list[dict[str, Any]]) -> None:
    header_row, cols = _find_item_header(ws)
    if header_row == 0:
        return

    row_idx = header_row + 1
    total_supply = 0
    total_tax = 0
    total_amount = 0
    for idx, item in enumerate(items, start=1):
        qty = _to_int(item.get("order_qty") or item.get("qty")) or _suggest_order_qty(item)
        unit_price = _to_int(item.get("unit_price") or item.get("paid_amount")) or 0
        supply = _to_int(item.get("supply_amount") or item.get("amount"))
        if supply is None:
            supply = qty * unit_price
        tax = _to_int(item.get("tax_amount"))
        if tax is None:
            tax = int(supply * 0.1)
        amount = supply + tax

        _set_cell(ws, row_idx, cols.get("no"), idx)
        _set_cell(ws, row_idx, cols.get("name"), item.get("prod_nm") or "")
        _set_cell(ws, row_idx, cols.get("spec"), item.get("spec") or "")
        _set_cell(ws, row_idx, cols.get("qty"), qty)
        _set_cell(ws, row_idx, cols.get("unit_price"), unit_price)
        _set_cell(ws, row_idx, cols.get("supply_amount") or cols.get("amount"), supply)
        _set_cell(ws, row_idx, cols.get("tax_amount"), tax)
        _set_cell(ws, row_idx, cols.get("note"), item.get("note") or "")

        total_supply += supply
        total_tax += tax
        total_amount += amount
        row_idx += 1

    _put_after_label(ws, {"합계", "합 계", "합      계"}, total_supply)
    _put_after_label(ws, {"발주금액", "총액", "합계금액"}, total_amount)


def _collect_items(req: AiRequest) -> list[dict[str, Any]]:
    raw_items = req.get_slot("approved_items") or req.get_slot("items") or []
    if not isinstance(raw_items, list):
        return []
    items: list[dict[str, Any]] = []
    for raw in raw_items:
        if not isinstance(raw, dict):
            continue
        if not str(raw.get("prod_nm") or "").strip():
            continue
        item = dict(raw)
        item["prod_stock"] = _to_int(raw.get("prod_stock"))
        item["order_qty"] = _to_int(raw.get("order_qty"))
        item["qty"] = _to_int(raw.get("qty"))
        item["paid_amount"] = _to_int(raw.get("paid_amount"))
        item["unit_price"] = _to_int(raw.get("unit_price"))
        item["amount"] = _to_int(raw.get("amount"))
        item["supply_amount"] = _to_int(raw.get("supply_amount"))
        item["tax_amount"] = _to_int(raw.get("tax_amount"))
        items.append(item)
    return items


def _find_item_header(ws: Worksheet) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row, 120) + 1):
        values = [str(ws.cell(row, col).value or "").replace(" ", "") for col in range(1, min(ws.max_column, 70) + 1)]
        if "품명" not in values:
            continue
        if "수량" not in values and "단가" not in values:
            continue
        mapping: dict[str, int] = {}
        for col, val in enumerate(values, start=1):
            if val in {"NO", "NO.", "번호"}:
                mapping["no"] = col
            elif val == "품명":
                mapping["name"] = col
            elif val == "규격":
                mapping["spec"] = col
            elif val == "수량":
                mapping["qty"] = col
            elif val == "단가":
                mapping["unit_price"] = col
            elif val == "공급가액":
                mapping["supply_amount"] = col
            elif val == "세액":
                mapping["tax_amount"] = col
            elif val == "금액":
                mapping["amount"] = col
            elif val == "비고":
                mapping["note"] = col
        return row, mapping
    return 0, {}


def _put_after_label(ws: Worksheet, labels: set[str], value: Any) -> None:
    normalized_labels = {label.replace(" ", "") for label in labels}
    for row in range(1, min(ws.max_row, 80) + 1):
        for col in range(1, min(ws.max_column, 80) + 1):
            cell_value = str(ws.cell(row, col).value or "").replace(" ", "")
            if cell_value not in normalized_labels:
                continue
            for step in range(1, 8):
                target_col = col + step
                if target_col > ws.max_column:
                    break
                target = ws.cell(row, target_col)
                # Skip merged/title labels and write to first writable cell.
                if step == 1 or target.value in (None, "", "(", ")"):
                    target.value = value
                    return


def _set_cell(ws: Worksheet, row: int, col: int | None, value: Any) -> None:
    if col is None:
        return
    ws.cell(row=row, column=col).value = value


def _suggest_order_qty(item: dict[str, Any]) -> int:
    stock = _to_int(item.get("prod_stock")) or 0
    target_stock = _to_int(item.get("target_stock")) or 12
    qty = target_stock - stock
    return qty if qty > 0 else 1


def _to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(str(value).replace(",", "").strip()))
        except (TypeError, ValueError):
            return None


def _safe(value: str) -> str:
    text = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in value.strip().lower())
    return text.strip("_") or "unknown"